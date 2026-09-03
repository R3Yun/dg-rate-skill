# -*- coding: utf-8 -*-
"""
Excel 文件直接读取与标准化
支持 .xlsx (openpyxl) 和 .xls (xlrd)
将任意 Excel 转换为「结构化原始数据」格式供后续 Parser 处理
"""
import os
import re
from typing import List, Dict, Any, Tuple
from pathlib import Path

try:
    import openpyxl
    from openpyxl.cell import MergedCell
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


class ExcelWorksheet:
    """工作表封装"""
    def __init__(self, name: str):
        self.name = name
        self.rows: List[List[str]] = []  # 所有单元格转为字符串
        self.merged_cells: List[Tuple[int, int, int, int]] = []  # (min_row, max_row, min_col, max_col), 0-based

    def __getitem__(self, idx):
        return self.rows[idx] if idx < len(self.rows) else []

    def __len__(self):
        return len(self.rows)

    def row_text(self, row_idx: int, join_str: str = " ") -> str:
        """获取某行所有单元格拼接文本"""
        if row_idx >= len(self.rows):
            return ""
        return join_str.join(str(c) for c in self.rows[row_idx] if c is not None and str(c).strip())

    def first_rows_text(self, n: int = 12, join_str: str = "\n") -> str:
        """获取前N行拼接文本（用于格式特征识别）"""
        return join_str.join(self.row_text(i) for i in range(min(n, len(self.rows))))

    def get_merged_value(self, row: int, col: int) -> str:
        """获取合并单元格的值（检查该单元格是否属于某个合并区域，返回左上角值）"""
        for (min_r, max_r, min_c, max_c) in self.merged_cells:
            if min_r <= row <= max_r and min_c <= col <= max_c:
                return str(self.rows[min_r][min_c]) if min_r < len(self.rows) and min_c < len(self.rows[min_r]) else ""
        return str(self.rows[row][col]) if row < len(self.rows) and col < len(self.rows[row]) else ""


class ExcelWorkbook:
    """工作簿封装"""
    def __init__(self, path: str):
        self.path = path
        self.sheets: Dict[str, ExcelWorksheet] = {}
        self._load()

    def _load(self):
        ext = os.path.splitext(self.path)[1].lower()
        if ext == ".xlsx" or ext == ".xlsm":
            self._load_openpyxl()
        elif ext == ".xls":
            self._load_xlrd()
        else:
            raise ValueError(f"不支持的Excel格式: {ext}")

    def _load_openpyxl(self):
        """用 openpyxl 读取 .xlsx"""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet = ExcelWorksheet(sheet_name)
            # 读取数据
            for row in ws.iter_rows(values_only=True):
                sheet.rows.append([str(c) if c is not None else "" for c in row])
            # 读取合并单元格
            for merged_range in ws.merged_cells.ranges:
                sheet.merged_cells.append((
                    merged_range.min_row - 1, merged_range.max_row - 1,
                    merged_range.min_col - 1, merged_range.max_col - 1
                ))
            self.sheets[sheet_name] = sheet
        wb.close()

    def _load_xlrd(self):
        """用 xlrd 读取 .xls"""
        if not HAS_XLRD:
            raise ImportError("xlrd 未安装，请执行: pip install xlrd")
        wb = xlrd.open_workbook(self.path)
        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet = ExcelWorksheet(ws.name)
            for row_idx in range(ws.nrows):
                row = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell_value(row_idx, col_idx)
                    row.append(str(cell) if cell is not None else "")
                sheet.rows.append(row)
            # 注意：xlrd 2.x 不支持 .merged_cells，跳过合并信息
            self.sheets[ws.name] = sheet

    @property
    def sheet_names(self) -> List[str]:
        return list(self.sheets.keys())

    def __getitem__(self, name) -> ExcelWorksheet:
        return self.sheets.get(name) or self.sheets.get(list(self.sheets.keys())[0])


def read_excel_to_text(path: str, sheet_name: str = None) -> str:
    """
    读取 Excel 并转为可被文本 Parser 处理的 TSV 格式文本
    这是最简单的兼容方式 - 让现有 TextParser 直接处理 Excel 内容
    """
    wb = ExcelWorkbook(path)
    if sheet_name and sheet_name in wb.sheets:
        ws = wb.sheets[sheet_name]
    else:
        ws = list(wb.sheets.values())[0]  # 默认第一个sheet
    # 转为 TSV 格式，方便现有 Parser 处理
    lines = []
    for row in ws.rows:
        clean_row = [str(c).replace("\t", " ").replace("\n", " ") for c in row]
        lines.append("\t".join(clean_row))
    return "\n".join(lines)


def excel_to_raw_structure(path: str) -> Dict[str, Any]:
    """
    Excel 转为「LLM 友好」的原始结构（供 inspect_excel.py 和 adaptive_transformer.py 使用）
    输出格式与 inspect_excel.py CSV 输出一致
    """
    wb = ExcelWorkbook(path)
    result = {
        "file": os.path.basename(path),
        "sheets": []
    }
    for sheet_name, ws in wb.sheets.items():
        data_rows = []
        for r_idx, row in enumerate(ws.rows[:50]):  # 前50行足够分析结构
            data_rows.append({
                "row_idx": r_idx,
                "cells": [str(c) for c in row[:30]]  # 前30列
            })
        result["sheets"].append({
            "name": sheet_name,
            "row_count": len(ws.rows),
            "merged_cells": ws.merged_cells,
            "first_12_rows_text": ws.first_rows_text(12),
            "data_rows": data_rows
        })
    return result


def detect_excel_parser(path: str) -> Tuple[str, float, str]:
    """
    从 Excel 文件特征自动检测应该用哪个 Parser
    返回 (parser_name, confidence, reason)
    """
    try:
        wb = ExcelWorkbook(path)
    except Exception as e:
        return ("", 0, f"Excel读取失败: {e}")

    scores = {
        "cargoware_template": 0.0,
        "sitc_notification": 0.0,
        "forwarder_summary": 0.0,
        "yml_fak": 0.0,
        "tier_guide": 0.0,
        "ial_tariff": 0.0,
        "cul_notice": 0.0,
        "msk_maersk_rate": 0.0,
        "cma_cgm_rate": 0.0,
        "evergreen_rate": 0.0,
        "hmm_rate": 0.0,
        "zim_rate": 0.0,
        "yang_ming_rate": 0.0,
        "hapag_lloyd_rate": 0.0,
        "one_line_rate": 0.0,
        "pil_rate": 0.0,
        "msc_rate": 0.0,
        "cosco_rate": 0.0,
    }

    filename = Path(path).name.lower()

    # 1. 检查是否是 CargoWare 模板
    # B1 (WS-149, 2026-09-01): 每格式每类特征只计一次 (any-sheet), 禁止跨 sheet 累加 —
    # 义统 11-sheet 文件每 sheet 都含 20'/40', 旧逻辑累加 11×0.2=2.2 淹没真实格式
    cargo_any_fcl31 = any("FCL3.1" in ws.first_rows_text(8) or "RATE TEMPLATE" in ws.first_rows_text(8)
                          for ws in wb.sheets.values())
    cargo_any_plpd = any("POL" in ws.first_rows_text(8) and "POD" in ws.first_rows_text(8)
                         and "CARRIER" in ws.first_rows_text(8) for ws in wb.sheets.values())
    cargo_any_quotes = any("20'" in ws.first_rows_text(8) or "40'" in ws.first_rows_text(8)
                           or "40'HC" in ws.first_rows_text(8) for ws in wb.sheets.values())
    if cargo_any_fcl31:
        scores["cargoware_template"] += 0.9
    if cargo_any_plpd:
        scores["cargoware_template"] += 0.3
    if cargo_any_quotes:
        scores["cargoware_template"] += 0.2

    # 2. 检查是否是 SITC 海丰格式
    if "海丰" in filename or "SITC" in filename.upper():
        scores["sitc_notification"] += 0.5
    sitc_any_head = any("SITC" in ws.first_rows_text(12) or "海丰" in ws.first_rows_text(12)
                        or "运价通知" in ws.first_rows_text(12) for ws in wb.sheets.values())
    if sitc_any_head:
        scores["sitc_notification"] += 0.5

    # 3. 检查是否是货代汇总表（义统等）
    # B1 (WS-149, 2026-09-01): 区域 sheet 名扩展为义统真实 11-sheet 命名 (澳洲/欧洲/地中海/
    # 东南亚/中东印巴/红海/南美/非洲/美国/订舱费/THC) + DESTINATION header 特征
    if "义统" in filename or "报价" in filename or "汇总" in filename:
        scores["forwarder_summary"] += 0.4
    fs_region_hits = 0
    fs_dest_hits = 0
    for sheet_name, ws in wb.sheets.items():
        if any(region in sheet_name for region in
               ("香港海防","泰越柬","韩国","日本","台湾","东南亚","澳洲","新西兰","欧洲","地中海",
                "中东","印巴","红海","南美","非洲","美国","北美","加勒比","朝鲜","俄罗斯","印度")):
            fs_region_hits += 1
        head_text = ws.first_rows_text(12)
        if "DESTINATION" in head_text and ("20'GP" in head_text or "20GP" in head_text or "20'" in head_text):
            fs_dest_hits += 1
        if "船公司" in head_text and "20\"/40\"" in head_text.replace("'","\""):
            scores["forwarder_summary"] += 0.4
    # 义统特征: ≥3 个区域 sheet + DESTINATION header → 强信号
    if fs_region_hits >= 3:
        scores["forwarder_summary"] += 0.8
    elif fs_region_hits >= 1:
        scores["forwarder_summary"] += 0.3
    if fs_dest_hits >= 1:
        scores["forwarder_summary"] += 0.5

    # 4. 检查 YML FAK TARIFF 格式
    if "yml" in filename or "fak" in filename or "tariff" in filename:
        scores["yml_fak"] += 0.4
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        if "FAK" in head_text and "Shipment kind" in head_text:
            scores["yml_fak"] += 0.6

    # 5. 检查 Tier Guide Rate 格式
    if "tier" in filename or "guide" in filename:
        scores["tier_guide"] += 0.5

    # 6. D51: 检查是否是 IAL 格式 (印巴/南亚航线指南)
    # B1 (WS-149, 2026-09-01): 每类特征只计一次 (any-sheet) — ETD 出现在任意表格 header
    # 都可能命中, 义统 11-sheet 文件每 sheet 都含 ETD, 旧逻辑累加淹没真实格式
    ial_any_name = any("IAL" in ws.first_rows_text(10) or "印度" in ws.first_rows_text(10)
                       or "巴基斯坦" in ws.first_rows_text(10) or "Sri Lanka" in ws.first_rows_text(10)
                       for ws in wb.sheets.values())
    ial_any_freight = any("Ocean Freight" in ws.first_rows_text(10) and "SVC" in ws.first_rows_text(10)
                          and "POD" in ws.first_rows_text(10) for ws in wb.sheets.values())
    ial_any_etd = any("SHA ETD" in ws.first_rows_text(10) or "ETD" in ws.first_rows_text(10)
                      for ws in wb.sheets.values())
    ial_any_tariff_sheet = any(sheet_name.lower() == "tariff" and "POD" in ws.first_rows_text(10)
                               for sheet_name, ws in wb.sheets.items())
    if ial_any_name:
        scores["ial_tariff"] += 0.3
    if ial_any_freight:
        scores["ial_tariff"] += 0.4
    if ial_any_etd:
        scores["ial_tariff"] += 0.2
    if ial_any_tariff_sheet:
        scores["ial_tariff"] += 0.2

    # 7. D51: 检查是否是 CUL 格式 (CHINA UNITED LINES 通知)
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        if "CHINA UNITED LINES" in head_text or "CUL" in head_text.upper()[:20]:
            scores["cul_notice"] += 0.6
        if "Selling Rate" in head_text and "航线" in head_text and "船名" in head_text:
            scores["cul_notice"] += 0.5
        if "20GP" in head_text and "40HC" in head_text and "ETD" in head_text:
            scores["cul_notice"] += 0.3

    # 8. D52: 检查是否是 MSK (MAERSK) 格式
    if "msk" in filename or "maersk" in filename:
        scores["msk_maersk_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        if "MAERSK" in head_text or "MSK" in head_text.upper()[:20]:
            scores["msk_maersk_rate"] += 0.5
        if "O/F(USD)" in head_text and "船公司" in head_text:
            scores["msk_maersk_rate"] += 0.3
        # D52 修复: 也检查数据行 (carrier 在数据行)
        data_text = ws.first_rows_text(20)
        if "MAERSK" in data_text or " MSK " in data_text:
            scores["msk_maersk_rate"] += 0.4

    # 9. D52: 检查是否是 CMA CGM 格式
    if "cma" in filename or "cgm" in filename:
        scores["cma_cgm_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        if "CMA CGM" in head_text or "CMA-CGM" in head_text:
            scores["cma_cgm_rate"] += 0.5
        if "O/F(USD)" in head_text and ("CMA" in head_text or "CGM" in head_text):
            scores["cma_cgm_rate"] += 0.3
        data_text = ws.first_rows_text(20)
        if "CMA CGM" in data_text or "CMA-CGM" in data_text:
            scores["cma_cgm_rate"] += 0.4

    # 10. D52: 检查是否是 EMC (Evergreen) 格式
    if "emc" in filename or "evergreen" in filename:
        scores["evergreen_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        if "EVERGREEN" in head_text or "EMC" in head_text.upper()[:20]:
            scores["evergreen_rate"] += 0.5
        if "O/F(USD)" in head_text and "船公司" in head_text:
            scores["evergreen_rate"] += 0.3
        data_text = ws.first_rows_text(20)
        if "EVERGREEN" in data_text or " EMC " in data_text:
            scores["evergreen_rate"] += 0.4

    # 11. D53: 检查是否是 HMM 格式
    if "hmm" in filename:
        scores["hmm_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        # D53 修复: 检查 head_text 全文 + data_text (carrier 在数据行)
        if "HMM" in head_text.upper():
            scores["hmm_rate"] += 0.5
        if "HMM" in data_text.upper():
            scores["hmm_rate"] += 0.5

    # 12. D53: 检查是否是 ZIM 格式
    if "zim" in filename:
        scores["zim_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        if "ZIM" in head_text.upper():
            scores["zim_rate"] += 0.5
        if "ZIM" in data_text.upper():
            scores["zim_rate"] += 0.5

    # 13. D53: 检查是否是 YANG MING 格式
    if "yang" in filename or "ming" in filename:
        scores["yang_ming_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        if "YANG MING" in head_text.upper():
            scores["yang_ming_rate"] += 0.6
        if "YANG MING" in data_text.upper():
            scores["yang_ming_rate"] += 0.6

    # 14. D53: 检查是否是 HAPAG-LLOYD 格式
    if "hapag" in filename or "lloyd" in filename:
        scores["hapag_lloyd_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        if "HAPAG" in head_text.upper() or "LLOYD" in head_text.upper():
            scores["hapag_lloyd_rate"] += 0.5
        if "HAPAG" in data_text.upper() or "LLOYD" in data_text.upper():
            scores["hapag_lloyd_rate"] += 0.5

    # 15. D53: 检查是否是 ONE (Ocean Network Express) 格式
    if "one" in filename:
        scores["one_line_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        if "ONE" in head_text.upper() and "OCEAN" in head_text.upper():
            scores["one_line_rate"] += 0.6
        elif "ONE" in head_text.upper() and len(head_text) < 100:
            # ONE 是 3 字符, 避免误判 (如 "ONE OF...")
            scores["one_line_rate"] += 0.3
        if "OCEAN NETWORK" in data_text.upper():
            scores["one_line_rate"] += 0.6
        elif "ONE" in data_text.upper() and "POL" in data_text.upper():
            scores["one_line_rate"] += 0.6  # D53 修复: 0.4 不够 0.5 阈值

    # 16. D54: 检查是否是 PIL 格式
    if "pil" in filename:
        scores["pil_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        if "PIL" in head_text.upper() and len(head_text) < 200:
            scores["pil_rate"] += 0.5
        if "PIL" in data_text.upper():
            scores["pil_rate"] += 0.5

    # 17. D54: 检查是否是 MSC 格式
    if "msc" in filename:
        scores["msc_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        if "MSC" in head_text.upper() and len(head_text) < 200:
            scores["msc_rate"] += 0.5
        if "MSC" in data_text.upper():
            scores["msc_rate"] += 0.5

    # 18. D54: 检查是否是 COSCO 格式
    if "cosco" in filename:
        scores["cosco_rate"] += 0.5
    for sheet_name, ws in wb.sheets.items():
        head_text = ws.first_rows_text(12)
        data_text = ws.first_rows_text(20)
        if "COSCO" in head_text.upper() and len(head_text) < 200:
            scores["cosco_rate"] += 0.5
        if "COSCO" in data_text.upper():
            scores["cosco_rate"] += 0.5

    best_parser = max(scores, key=scores.get)
    best_score = scores[best_parser]

    if best_score >= 0.5:
        return (best_parser, best_score, f"Excel特征匹配: {best_parser}")
    return ("adaptive", 0.3, "未匹配已知格式，建议用LLM-Adaptive模式")


if __name__ == "__main__":
    import sys
    import json
    if len(sys.argv) > 1:
        struct = excel_to_raw_structure(sys.argv[1])
        print(json.dumps(struct, ensure_ascii=False, indent=2))
    else:
        print("Usage: python excel_helper.py <file.xlsx>")

