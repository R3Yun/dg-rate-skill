#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Full XLS/XLSX extraction into a persistent ParseWorkspace."""

from __future__ import annotations

import json
import math
import os
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from parse_workspace import ParseWorkspace


def _json_value(value: Any) -> Tuple[Any, str]:
    if value is None:
        return None, "empty"
    if isinstance(value, datetime):
        return value.isoformat(), "datetime"
    if isinstance(value, date):
        return value.isoformat(), "date"
    if isinstance(value, time):
        return value.isoformat(), "time"
    if isinstance(value, bool):
        return value, "boolean"
    if isinstance(value, int):
        return value, "number"
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value), "text"
        return value, "number"
    return str(value), "text"


def _display_value(value: Any) -> str:
    normalized, _ = _json_value(value)
    if normalized is None:
        return ""
    if isinstance(normalized, bool):
        return "TRUE" if normalized else "FALSE"
    return str(normalized)


def _column_letter(index: int) -> str:
    letters = ""
    value = index
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _sheet_id(index: int) -> str:
    return f"sheet-{index:03d}"


def _safe_sheet_filename(sheet_id: str) -> str:
    return f"raw/{sheet_id}.jsonl"


def _markdown_escape(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("\r", " ").replace("\n", " ")


def rows_to_markdown(sheet_name: str, rows: Sequence[Dict[str, Any]], max_columns: int = 30) -> str:
    if not rows:
        return f"## Sheet: {sheet_name}\n\n（无行）"
    max_col = 0
    for row in rows:
        for cell in row.get("cells", []):
            max_col = max(max_col, int(cell.get("column_index", 0) or 0))
    max_col = min(max_col, max_columns)
    headers = ["Excel行"] + [_column_letter(i) for i in range(1, max_col + 1)]
    lines = [f"## Sheet: {sheet_name}", "", "| " + " | ".join(headers) + " |"]
    lines.append("|" + "|".join(["---"] * len(headers)) + "|")
    for row in rows:
        values = {int(c["column_index"]): c.get("display", "") for c in row.get("cells", [])}
        rendered = [str(row.get("row_number", ""))]
        rendered.extend(_markdown_escape(values.get(i, "")) for i in range(1, max_col + 1))
        lines.append("| " + " | ".join(rendered) + " |")
    if max_col < max(
        [int(c.get("column_index", 0) or 0) for row in rows for c in row.get("cells", [])] or [0]
    ):
        lines.append("")
        lines.append(f"> Markdown 视图仅显示前 {max_columns} 列；完整列已保存在 JSONL。")
    return "\n".join(lines)


def _write_sheet_rows(workspace: ParseWorkspace, parse_id: str, sheet_id: str, rows: Iterable[Dict[str, Any]]) -> None:
    target = workspace._safe_artifact_path(parse_id, _safe_sheet_filename(sheet_id))
    temporary = target.parent / f".{target.name}.tmp"
    try:
        with temporary.open("w", encoding="utf-8", newline="\n") as output:
            for row in rows:
                output.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
            output.flush()
            os.fsync(output.fileno())
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()


def _summary_markdown(source_file: str, sheets: Sequence[Dict[str, Any]]) -> str:
    lines = [f"# 运价文件解析摘要：{source_file}", "", f"- Sheet 数：{len(sheets)}", "", "| Sheet | 总行 | 非空行 | 总列 | 隐藏 |", "|---|---:|---:|---:|---|"]
    for sheet in sheets:
        lines.append(
            "| {name} | {total_rows} | {non_empty_rows} | {total_columns} | {hidden} |".format(
                name=_markdown_escape(sheet["name"]),
                total_rows=sheet["total_rows"],
                non_empty_rows=sheet["non_empty_rows"],
                total_columns=sheet["total_columns"],
                hidden="是" if sheet.get("hidden") else "否",
            )
        )
    return "\n".join(lines) + "\n"


def _create_workspace(source_path: str, root: Optional[str], chat_id: str, message_id: str) -> Tuple[ParseWorkspace, Dict[str, Any]]:
    workspace = ParseWorkspace(root)
    created = workspace.create(source_path, chat_id=chat_id, message_id=message_id)
    return workspace, created


def extract_xlsx(
    source_path: str,
    *,
    root: Optional[str] = None,
    chat_id: str = "",
    message_id: str = "",
    sheet_name: Optional[str] = None,
    sample_rows: int = 10,
) -> Dict[str, Any]:
    import openpyxl

    source = Path(source_path).expanduser().resolve()
    workbook_formula = openpyxl.load_workbook(source, data_only=False, read_only=False)
    workbook_values = openpyxl.load_workbook(source, data_only=True, read_only=False)
    selected = [ws for ws in workbook_formula.worksheets if not sheet_name or ws.title == sheet_name]
    if sheet_name and not selected:
        raise ValueError(f"sheet not found: {sheet_name}")

    workspace, created = _create_workspace(str(source), root, chat_id, message_id)
    parse_id = created["manifest"]["parse_id"]
    sheet_summaries = []
    preview_parts = []
    try:
        for index, ws in enumerate(selected, 1):
            values_ws = workbook_values[ws.title]
            current_sheet_id = _sheet_id(index)
            max_row = int(ws.max_row or 0)
            max_column = int(ws.max_column or 0)
            hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}
            hidden_columns = {key for key, dim in ws.column_dimensions.items() if dim.hidden}
            # 合并单元格: openpyxl 对合并区域内非左上角单元格返回 None, 需把左上角的
            # 原始值/显示值/是否公式传播到区域内每个 (row, col), 避免"备注只写第一条"。
            merged_lookup: Dict[Tuple[int, int], Tuple[Any, Any, bool]] = {}
            for rng in ws.merged_cells.ranges:
                tl_row, tl_col = rng.min_row, rng.min_col
                formula_tl = ws.cell(row=tl_row, column=tl_col)
                values_tl = values_ws.cell(row=tl_row, column=tl_col)
                tl_raw = formula_tl.value
                tl_is_formula = formula_tl.data_type == "f"
                tl_display_source = values_tl.value if tl_is_formula else tl_raw
                if tl_raw is None and tl_display_source is None:
                    continue  # 左上角本身为空, 无需传播
                for r in range(rng.min_row, rng.max_row + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        if r == tl_row and c == tl_col:
                            continue
                        merged_lookup[(r, c)] = (tl_raw, tl_display_source, tl_is_formula)
            preview_rows = []
            non_empty_rows = 0

            def iter_rows():
                nonlocal non_empty_rows
                for row_number in range(1, max_row + 1):
                    cells = []
                    is_empty = True
                    for column_index in range(1, max_column + 1):
                        formula_cell = ws.cell(row=row_number, column=column_index)
                        display_cell = values_ws.cell(row=row_number, column=column_index)
                        raw_value = formula_cell.value
                        is_formula = formula_cell.data_type == "f"
                        display_source = display_cell.value if is_formula else raw_value
                        # 合并单元格传播: 非左上角且自身值为空时用左上角值,
                        # 修复 openpyxl 对合并区域返回 None 导致"备注只写第一条"
                        merge_entry = merged_lookup.get((row_number, column_index))
                        if merge_entry is not None and raw_value is None and display_source is None:
                            raw_value, display_source, is_formula = merge_entry
                        json_raw, value_type = _json_value(raw_value)
                        json_display, display_type = _json_value(display_source)
                        if json_raw is not None or json_display is not None:
                            is_empty = False
                        column = _column_letter(column_index)
                        cells.append({
                            "column": column,
                            "column_index": column_index,
                            "coordinate": f"{column}{row_number}",
                            "raw": json_raw,
                            "display": _display_value(display_source),
                            "display_value": json_display,
                            "data_type": "formula" if is_formula else value_type,
                            "display_type": display_type,
                            "is_formula": is_formula,
                            "is_hidden": column in hidden_columns,
                        })
                    if not is_empty:
                        non_empty_rows += 1
                    row_payload = {
                        "source_row_id": f"{current_sheet_id}!{row_number}",
                        "sheet_id": current_sheet_id,
                        "sheet_name": ws.title,
                        "row_number": row_number,
                        "is_hidden": row_number in hidden_rows,
                        "is_empty": is_empty,
                        "cells": cells,
                    }
                    if len(preview_rows) < max(0, sample_rows):
                        preview_rows.append(row_payload)
                    yield row_payload

            _write_sheet_rows(workspace, parse_id, current_sheet_id, iter_rows())
            summary = {
                "sheet_id": current_sheet_id,
                "name": ws.title,
                "total_rows": max_row,
                "non_empty_rows": non_empty_rows,
                "total_columns": max_column,
                "hidden": ws.sheet_state != "visible",
                "hidden_rows": len(hidden_rows),
                "hidden_columns": len(hidden_columns),
                "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
                "raw_file": _safe_sheet_filename(current_sheet_id),
            }
            sheet_summaries.append(summary)
            preview_parts.append(rows_to_markdown(ws.title, preview_rows))

        workbook_meta = {
            "schema_version": "rate-workbook-raw/v1",
            "parse_id": parse_id,
            "parser": "read-xlsx-v2",
            "source_file": source.name,
            "sheet_count": len(sheet_summaries),
            "sheets": sheet_summaries,
        }
        summary_markdown = _summary_markdown(source.name, sheet_summaries)
        workspace.write_json(parse_id, "raw/workbook.json", workbook_meta)
        workspace.write_text(parse_id, "preview/summary.md", summary_markdown)
        state = workspace.update_state(
            parse_id,
            expected_revision=1,
            status="mapping_required",
            phase="extraction",
            last_action="workbook_extracted",
            next_action="inspect_sheet_pages",
            updates={
                "sheet_count": len(sheet_summaries),
                "total_rows": sum(sheet["total_rows"] for sheet in sheet_summaries),
                "non_empty_rows": sum(sheet["non_empty_rows"] for sheet in sheet_summaries),
            },
        )
        return _result_payload(created, state, workbook_meta, summary_markdown, preview_parts, sample_rows)
    except Exception:
        try:
            workspace.update_state(
                parse_id,
                expected_revision=workspace.load(parse_id)["state"]["revision"],
                status="failed_recoverable",
                phase="extraction",
                last_action="workbook_extraction_failed",
                next_action="retry_extraction",
            )
        except Exception:
            pass
        raise
    finally:
        workbook_formula.close()
        workbook_values.close()


def extract_xls(
    source_path: str,
    *,
    root: Optional[str] = None,
    chat_id: str = "",
    message_id: str = "",
    sheet_name: Optional[str] = None,
    sample_rows: int = 10,
) -> Dict[str, Any]:
    import xlrd

    source = Path(source_path).expanduser().resolve()
    workbook = xlrd.open_workbook(str(source), formatting_info=False)
    selected = [workbook.sheet_by_index(i) for i in range(workbook.nsheets) if not sheet_name or workbook.sheet_by_index(i).name == sheet_name]
    if sheet_name and not selected:
        raise ValueError(f"sheet not found: {sheet_name}")

    workspace, created = _create_workspace(str(source), root, chat_id, message_id)
    parse_id = created["manifest"]["parse_id"]
    sheet_summaries = []
    preview_parts = []
    for index, sheet in enumerate(selected, 1):
        current_sheet_id = _sheet_id(index)
        preview_rows = []
        non_empty_rows = 0

        def iter_rows():
            nonlocal non_empty_rows
            for row_index in range(sheet.nrows):
                row_number = row_index + 1
                cells = []
                is_empty = True
                for column_index_zero in range(sheet.ncols):
                    column_index = column_index_zero + 1
                    cell = sheet.cell(row_index, column_index_zero)
                    raw_value = cell.value
                    value_type = "text"
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        raw_value = None
                        value_type = "empty"
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        value_type = "number"
                        if isinstance(raw_value, float) and raw_value.is_integer():
                            raw_value = int(raw_value)
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        parts = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
                        if parts[0:3] == (0, 0, 0):
                            raw_value = time(parts[3], parts[4], parts[5]).isoformat()
                            value_type = "time"
                        else:
                            raw_value = datetime(*parts).isoformat()
                            value_type = "datetime"
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        raw_value = bool(raw_value)
                        value_type = "boolean"
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        raw_value = xlrd.error_text_from_code.get(cell.value, str(cell.value))
                        value_type = "error"
                    else:
                        raw_value = str(raw_value)
                    if raw_value not in (None, ""):
                        is_empty = False
                    column = _column_letter(column_index)
                    cells.append({
                        "column": column,
                        "column_index": column_index,
                        "coordinate": f"{column}{row_number}",
                        "raw": raw_value,
                        "display": _display_value(raw_value),
                        "display_value": raw_value,
                        "data_type": value_type,
                        "display_type": value_type,
                        "is_formula": False,
                        "is_hidden": False,
                    })
                if not is_empty:
                    non_empty_rows += 1
                payload = {
                    "source_row_id": f"{current_sheet_id}!{row_number}",
                    "sheet_id": current_sheet_id,
                    "sheet_name": sheet.name,
                    "row_number": row_number,
                    "is_hidden": False,
                    "is_empty": is_empty,
                    "cells": cells,
                }
                if len(preview_rows) < max(0, sample_rows):
                    preview_rows.append(payload)
                yield payload

        _write_sheet_rows(workspace, parse_id, current_sheet_id, iter_rows())
        summary = {
            "sheet_id": current_sheet_id,
            "name": sheet.name,
            "total_rows": sheet.nrows,
            "non_empty_rows": non_empty_rows,
            "total_columns": sheet.ncols,
            "hidden": False,
            "hidden_rows": 0,
            "hidden_columns": 0,
            "merged_ranges": [
                f"{_column_letter(c_lo + 1)}{r_lo + 1}:{_column_letter(c_hi)}{r_hi}"
                for r_lo, r_hi, c_lo, c_hi in sheet.merged_cells
            ],
            "raw_file": _safe_sheet_filename(current_sheet_id),
            "unsupported_metadata": ["hidden_rows", "hidden_columns", "formula_text"],
        }
        sheet_summaries.append(summary)
        preview_parts.append(rows_to_markdown(sheet.name, preview_rows))

    workbook_meta = {
        "schema_version": "rate-workbook-raw/v1",
        "parse_id": parse_id,
        "parser": "read-xls-v2",
        "source_file": source.name,
        "sheet_count": len(sheet_summaries),
        "sheets": sheet_summaries,
    }
    summary_markdown = _summary_markdown(source.name, sheet_summaries)
    workspace.write_json(parse_id, "raw/workbook.json", workbook_meta)
    workspace.write_text(parse_id, "preview/summary.md", summary_markdown)
    state = workspace.update_state(
        parse_id,
        expected_revision=1,
        status="mapping_required",
        phase="extraction",
        last_action="workbook_extracted",
        next_action="inspect_sheet_pages",
        updates={
            "sheet_count": len(sheet_summaries),
            "total_rows": sum(sheet["total_rows"] for sheet in sheet_summaries),
            "non_empty_rows": sum(sheet["non_empty_rows"] for sheet in sheet_summaries),
        },
    )
    return _result_payload(created, state, workbook_meta, summary_markdown, preview_parts, sample_rows)


def _result_payload(created, state, workbook_meta, summary_markdown, preview_parts, sample_rows):
    manifest = created["manifest"]
    return {
        "code": "PARSE_CREATED",
        "parse_id": manifest["parse_id"],
        "state": state["status"],
        "revision": state["revision"],
        "source_summary": {
            "file": manifest["source_file"],
            "size_kb": round(manifest["source_size"] / 1024, 1),
            "sha256": manifest["source_sha256"],
            "sheet_count": workbook_meta["sheet_count"],
            "sheets": workbook_meta["sheets"],
            "parser": workbook_meta["parser"].replace("-v2", ""),
            "workspace_parser": workbook_meta["parser"],
        },
        "content_markdown": summary_markdown + "\n" + "\n\n".join(preview_parts),
        "reading_hint": (
            f"完整工作簿已保存到 parse_id={manifest['parse_id']}；"
            f"当前仅返回每个 Sheet 前 {sample_rows} 行样本，后续用 rate-parse-page 分页读取。"
        ),
    }