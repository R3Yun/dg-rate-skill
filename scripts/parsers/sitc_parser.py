import os
import re
from typing import List
from .base import BaseRateParser, register_parser, NormalizedRateEntry

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class SitcParser(BaseRateParser):
    name = "sitc_haifeng"
    source_type = "sitc_haifeng"
    VESSEL_RE = re.compile(r"\b([A-Z]{2,4}\d+)\b")
    # Skip rows containing these patterns (fee/remark rows)
    SKIP_PATTERNS = [
        re.compile(r"冻桌|超重费|丢船费|文件费|港费|RCS|BAF|CAF|LSS|ETD|DAD|ICD|拆盖|日费|月费|报关|清关"),
        re.compile(r"加收|收取标准|标准收费|费用说明|提醒|注意"),
        re.compile(r"^\d+[、.]"),  # starts with number + Chinese comma
        re.compile(r"\d{4}/\d{1,2}/\d{1,2}"),  # date patterns like 2014.4.15
    ]

    def can_parse(self, content, filename="") -> tuple:
        score = 0.0; reason = []
        if filename.lower().endswith(".xlsx") and HAS_OPENPYXL:
            score += 0.5; reason.append("xlsx")
        for kw in ["海丰", "SITC", "sitc", "haifeng"]:
            if kw in filename:
                score += 0.4; reason.append(kw); break
        return (min(score, 1.0), "; ".join(reason))

    def parse(self, content="", filename="") -> List[NormalizedRateEntry]:
        # filename 可能是 "xxx.xlsx" 或 "xxx.xlsx[SheetName]"（parse_excel_file 传入）
        fp = filename.split("[")[0]
        if fp.lower().endswith(".xlsx") and os.path.exists(fp):
            return self._parse_xlsx(fp)
        return []

    def _parse_xlsx(self, fp) -> List[NormalizedRateEntry]:
        if not HAS_OPENPYXL: raise RuntimeError("openpyxl not installed")
        results = []
        wb = openpyxl.load_workbook(fp, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = self._get_rows_with_merges(ws)
            results.extend(self._parse_sheet(rows, fp, sname))
        wb.close()
        return results

    def _get_rows_with_merges(self, ws):
        merge_map = {}
        for mr in ws.merged_cells.ranges:
            mc = mr.bounds
            val = ws.cell(mc[1], mc[0]).value
            val_str = str(val).strip() if val is not None else ""
            for r in range(mc[1], mc[3]+1):
                for c in range(mc[0], mc[2]+1):
                    merge_map[(r, c)] = val_str

        result = []
        for row in ws.iter_rows(values_only=False):
            row_vals = []
            for cell in row:
                key = (cell.row, cell.column)
                if key in merge_map:
                    row_vals.append(merge_map[key])
                else:
                    v = cell.value
                    row_vals.append(str(v).strip() if v is not None else "")
            result.append(row_vals)
        return result

    def _is_fee_row(self, pod: str, price: str) -> bool:
        """Check if a row is a fee/remark row rather than a shipping rate."""
        text = (pod + " " + price).strip()
        if not text: return True
        for pat in self.SKIP_PATTERNS:
            if pat.search(text): return True
        return False

    def _is_valid_port(self, text: str) -> bool:
        """Check if text looks like a port name (not a fee/note)."""
        if not text: return False
        # Must not start with digit + comma pattern
        if re.match(r"^\d+[u3001.,]", text): return False
        # Must not contain common fee keywords
        if self._is_fee_row(text, ""): return False
        # Must have some meaningful length
        if len(text) < 2: return False
        return True

    def _parse_sheet(self, rows, fp, sheet_name) -> List[NormalizedRateEntry]:
        results = []
        schedule = ""
        for ri, row in enumerate(rows):
            if ri < 7: continue
            if len(row) < 5: continue

            hx = row[0].strip() if len(row) > 0 else ""
            bj = row[1].strip() if len(row) > 1 else ""
            gk = row[2].strip() if len(row) > 2 else ""
            yj = row[4].strip() if len(row) > 4 else ""
            rq = row[8].strip() if len(row) > 8 else ""
            bz = row[10].strip() if len(row) > 10 else ""

            if not any([hx, bj, gk, yj, rq]): continue

            has_vessel = bool(self.VESSEL_RE.search(bj)) if bj else False

            if has_vessel:
                freq = self._extract_frequency(bj)
                if freq: schedule = freq
                if gk and self._is_valid_port(gk):
                    entry = self._make_entry(gk, yj, rq, bz, fp, sheet_name, schedule)
                    if entry: results.append(entry)
            elif gk and self._is_valid_port(gk):
                entry = self._make_entry(gk, yj, rq, bz, fp, sheet_name, schedule)
                if entry: results.append(entry)

        return results

    def _extract_frequency(self, text: str) -> str:
        if not text: return ""
        day_map = {"周一":"MON","周二":"TUE","周三":"WED",
                    "周四":"THU","周五":"FRI","周六":"SAT","周日":"SUN"}
        vessel = self.VESSEL_RE.search(text)
        vessel_code = vessel.group(1) if vessel else ""
        for cn, en in day_map.items():
            if cn in text: return en + ("/" + vessel_code if vessel_code else "")
        return vessel_code

    def _make_entry(self, pod, price_text, date_text, remark, fp, sheet, frequency):
        if not pod: return None
        prices = self._parse_prices(price_text)
        if not any(prices.values()): return None
        entry = self.new_entry(fp)
        entry.pol = "CNSHA"
        entry.pol_name = "上海"
        entry.pod = self.normalize_port(pod)
        entry.pod_name = pod
        entry.carrier = "SITC"
        entry.carrier_name = "新海丰物流"
        entry.of_20 = prices.get("20")
        entry.of_40 = prices.get("40")
        entry.of_40hq = prices.get("40HC")
        entry.frequency = frequency
        entry.valid_from = self._parse_date(date_text)
        entry.remark = remark
        entry.source_type = self.source_type
        if not entry.pod:
            entry.warnings.append("Unknown port: " + pod)
        return entry

    def _parse_prices(self, text: str):
        result = {"20": None, "40": None, "40HC": None}
        if not text: return result
        amounts = re.findall(r"\$(\d+)", text)
        if len(amounts) >= 3:
            result["20"] = float(amounts[0])
            result["40"] = float(amounts[1])
            result["40HC"] = float(amounts[2])
        elif len(amounts) == 2:
            result["20"] = float(amounts[0])
            result["40"] = float(amounts[1])
        elif len(amounts) == 1:
            result["20"] = float(amounts[0])
            result["40"] = float(amounts[0])
        return result

    def _parse_date(self, text: str) -> str:
        if not text: return ""
        text = str(text).strip()
        m = re.match(r"(\d{4})-(\d{2})-(\d{2})", text)
        if m: return m.group(1) + "-" + m.group(2) + "-" + m.group(3)
        return text

register_parser(SitcParser())
