# -*- coding: utf-8 -*-
"""CargoWare FCL3.1 运价模板解析器
支持: .xlsx(openpyxl), .xls(xlrd), CSV/TSV文本(tab/逗号分隔)
前6行元数据，第7行列名，第8行起数据。
"""
import os, re, sys, csv, io
from typing import List, Tuple, Optional
from .base import BaseRateParser, register_parser, NormalizedRateEntry, DGSurcharge

try:
    import openpyxl; HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd; HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

class CwTemplateParser(BaseRateParser):
    name = "cargoware_fcl_template"
    source_type = "cargoware_template"

    COLUMN_PATTERNS = {
        "pol": re.compile(r"^\s*POL\s*$", re.I),
        "pod": re.compile(r"^\s*POD\s*$", re.I),
        "via_port": re.compile(r"^\s*VIA\s*$", re.I),
        "direct": re.compile(r"^\s*DIRECT\s*$", re.I),
        "tt_days": re.compile(r"^\s*T/?T\s*$", re.I),
        "frequency": re.compile(r"^\s*FREQ", re.I),
        "carrier": re.compile(r"^\s*CARRIER", re.I),
        "booking_agent": re.compile(r"^\s*BOOKING\s*AGENT", re.I),
        "of_20": re.compile(r"^\s*20['\"]?\s*$", re.I),
        "of_40": re.compile(r"^\s*40['\"]?\s*$", re.I),
        "of_40hq": re.compile(r"^\s*40['\"]?\s*HC", re.I),
        "of_20nor": re.compile(r"^\s*20['\"]?\s*NOR", re.I),
        "of_40nor": re.compile(r"^\s*40['\"]?\s*NOR", re.I),
        "of_45": re.compile(r"^\s*45['\"]?\s*$", re.I),
        "valid_from": re.compile(r"^\s*VALID\s*(FROM|FM|DATE\s*FM)", re.I),
        "valid_to": re.compile(r"^\s*VALID\s*(TO|TILL)", re.I),
        "remark": re.compile(r"^\s*REMARK", re.I),
        "contract_no": re.compile(r"^\s*CONTRACT", re.I),
        "ens": re.compile(r"^\s*ENS", re.I),
        "ams": re.compile(r"^\s*AMS", re.I),
        "ows_note": re.compile(r"^\s*(OVERLOAD|OW)", re.I),
        "free_time": re.compile(r"^\s*FREE\s*TIME", re.I),
    }

    def can_parse(self, content, filename="") -> Tuple[float, str]:
        score, reason = 0.0, []
        ext = os.path.splitext(filename)[1].lower()
        if ext in (".xlsx", ".xls"):
            if ext == ".xlsx" and HAS_OPENPYXL: score += 0.9; reason.append("xlsx(openpyxl)")
            elif ext == ".xls" and HAS_XLRD: score += 0.9; reason.append("xls(xlrd)")
            else: score += 0.2; reason.append("Excel(库未装)")
        if not content: return (score, "; ".join(reason))
        fn = filename.upper()
        if "FCL" in fn or "CW" in fn: score += 0.3; reason.append("文件名")
        lines = content.splitlines()
        if len(lines) >= 7:
            head = "\n".join(lines[:7]).upper()
            if "FCL3.1" in head or "POL" in head: score += 0.4; reason.append("含列名")
        return (min(score, 1.0), "; ".join(reason))

    def parse(self, content="", filename="") -> List[NormalizedRateEntry]:
        ext = os.path.splitext(filename)[1].lower()
        if ext == ".xlsx": return self._parse_xlsx(filename)
        if ext == ".xls": return self._parse_xls(filename)
        return self._parse_text(content, filename)

    def _parse_xlsx(self, fp) -> List[NormalizedRateEntry]:
        if not HAS_OPENPYXL: raise RuntimeError("openpyxl未安装")
        results = []
        wb = openpyxl.load_workbook(fp, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            results.extend(self._parse_rows(rows, fp))
        wb.close()
        return results

    def _parse_xls(self, fp) -> List[NormalizedRateEntry]:
        if not HAS_XLRD: raise RuntimeError("xlrd未安装")
        results = []
        wb = xlrd.open_workbook(fp)
        for si in range(wb.nsheets):
            ws = wb.sheet_by_index(si)
            rows = [[ws.cell_value(ri, ci) for ci in range(ws.ncols)] for ri in range(ws.nrows)]
            results.extend(self._parse_rows(rows, fp))
        return results

    def _parse_rows(self, rows, fp) -> List[NormalizedRateEntry]:
        results, col_map, header_idx = [], {}, -1
        for ri, row in enumerate(rows):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if not any(vals): continue
            pol_cnt = sum(1 for v in vals if re.match(r"^\s*POL\s*$", v, re.I))
            if pol_cnt > 0:
                header_idx = ri
                for ci, v in enumerate(vals):
                    for f, pat in self.COLUMN_PATTERNS.items():
                        if pat.search(v) and f not in col_map:
                            col_map[f] = ci
                break
        if not col_map: return results
        currency = "USD"
        for ri in range(min(6, len(rows))):
            for v in rows[ri]:
                if str(v).strip().upper() in ("USD","RMB","CNY","EUR"):
                    currency = str(v).strip().upper()
        for ri in range(header_idx + 1, len(rows)):
            row = [str(v).strip() if v is not None else "" for v in rows[ri]]
            first = row[0].strip() if row else ""
            if first not in ("+","*","") and not re.match(r"^\d+$", first): continue
            entry = self._parse_row(row, col_map, currency, fp)
            if entry: results.append(entry)
        return results

    def _parse_text(self, content, filename) -> List[NormalizedRateEntry]:
        results, col_map, header_idx = [], {}, -1
        lines = content.splitlines()
        if not lines: return results
        currency = "USD"
        for i in range(min(6, len(lines))):
            vals = self._split_row(lines[i])
            for v in vals:
                if v.strip().upper() in ("USD","RMB","CNY","EUR"):
                    currency = v.strip().upper()
        for ri, line in enumerate(lines):
            row = self._split_row(line)
            vals = [v.strip() for v in row]
            if not any(vals): continue
            pol_cnt = sum(1 for v in vals if re.match(r"^\s*POL\s*$", v, re.I))
            if pol_cnt > 0:
                header_idx = ri
                for ci, v in enumerate(vals):
                    for f, pat in self.COLUMN_PATTERNS.items():
                        if pat.search(v) and f not in col_map:
                            col_map[f] = ci
                break
        if not col_map: return results
        for line in lines[header_idx + 1:]:
            row = self._split_row(line)
            vals = [v.strip() for v in row]
            first = vals[0].strip() if vals else ""
            if first not in ("+","*","") and not re.match(r"^\d+$", first): continue
            entry = self._parse_row(vals, col_map, currency, filename)
            if entry: results.append(entry)
        return results

    def _split_row(self, line):
        if "\t" in line: return [v.strip() for v in line.split("\t")]
        try: return next(csv.reader(io.StringIO(line)))
        except: pass
        return [v.strip() for v in line.split(",")]

    def _cell(self, row, key):
        idx = self._col_idx.get(key, -1)
        return row[idx].strip() if 0 <= idx < len(row) else ""

    def _parse_row(self, row, col_map, currency, filename) -> Optional[NormalizedRateEntry]:
        self._col_idx = col_map
        e = self.new_entry(filename)
        e.currency = currency
        e.pol = self._cell(row, "pol")
        e.pod = self._cell(row, "pod")
        # 保留原始港口名到 *_name, 用 normalize_port 标准化代码
        e.pol_name = e.pol
        e.pod_name = e.pod
        e.pol = self.normalize_port(e.pol)
        e.pod = self.normalize_port(e.pod)
        e.via_port = self._cell(row, "via_port")
        e.carrier = self._cell(row, "carrier")
        e.booking_agent = self._cell(row, "booking_agent")
        e.contract_no = self._cell(row, "contract_no")
        e.of_20 = self.parse_price(self._cell(row, "of_20"))
        e.of_40 = self.parse_price(self._cell(row, "of_40"))
        e.of_40hq = self.parse_price(self._cell(row, "of_40hq"))
        e.of_20nor = self.parse_price(self._cell(row, "of_20nor"))
        e.of_40nor = self.parse_price(self._cell(row, "of_40nor"))
        e.of_45 = self.parse_price(self._cell(row, "of_45"))
        tt = self._cell(row, "tt_days")
        e.tt_days = int(tt) if tt.isdigit() else None
        e.frequency = self._cell(row, "frequency")
        d = self._cell(row, "direct").upper()
        e.direct = "Y" if ("Y" in d or "直" in d) else ("T" if ("T" in d or "中" in d) else "")
        vf = self._cell(row, "valid_from"); vt = self._cell(row, "valid_to")
        e.valid_from = self._fmt_date(vf)
        e.valid_to = self._fmt_date(vt)
        remark = self._cell(row, "remark")
        if remark:
            e.remark = remark
            e.dg_surcharges = self.parse_dg_remark(remark)
            if not e.valid_from or not e.valid_to:
                rvf, rvt = self.parse_date(remark)
                e.valid_from = e.valid_from or rvf; e.valid_to = e.valid_to or rvt
        e.ens = self.parse_price(self._cell(row, "ens"))
        e.ams = self.parse_price(self._cell(row, "ams"))
        e.ows_note = self._cell(row, "ows_note")
        e.raw_excerpt = "|".join(str(v) for v in row[:10])
        if not e.pol or not e.pod:
            e.warnings.append("缺港口"); return None
        if not e.of_20 and not e.of_40:
            e.warnings.append("无价格"); return None
        return e

    def _fmt_date(self, s):
        if not s: return ""
        s = str(s).strip()
        if re.match(r"^\d{4,5}$", s):
            try:
                import datetime
                serial = int(s)
                if serial > 50000 and HAS_XLRD:
                    d = datetime.datetime(*xlrd.xldate_as_tuple(serial, 0))
                else:
                    d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=serial)
                return d.strftime("%Y-%m-%d")
            except: pass
        m = re.match(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", s)
        if m: return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        return s

register_parser(CwTemplateParser())